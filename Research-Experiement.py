import networkx as nx
import numpy as np
from networkx.algorithms.coloring import greedy_color
import matplotlib.pyplot as plt
import pandas as pd  
import scipy.stats as stats  
from sklearn.linear_model import LinearRegression 
from openpyxl import Workbook
from openpyxl.drawing.image import Image 

def calculate_network_metrics(G):
    """
    Calculate all network metrics in one place
    """
    # Calculate spectral gap
    laplacian = nx.laplacian_matrix(G).toarray()
    eigenvalues = np.linalg.eigvalsh(laplacian)
    eigenvalues.sort()
    spectral_gap = eigenvalues[1]
    
    # Perform coloring
    coloring = nx.greedy_color(G)
    num_colors = len(set(coloring.values()))
    
    # Calculate additional metrics
    metrics = {
        'spectral_gap': spectral_gap,
        'num_colors': num_colors,
        'clustering': nx.average_clustering(G),
        'avg_path_length': nx.average_shortest_path_length(G) if nx.is_connected(G) else np.nan,
        'diameter': nx.diameter(G) if nx.is_connected(G) else np.nan
    }
    
    return metrics

def generate_and_analyze_network(n, k, p):
    """
    Generate a WS network and analyze its properties
    """
    # Set a different random seed for each instance
    np.random.seed(None)  # Ensures different seed each time
    G = nx.watts_strogatz_graph(n=n, k=k, p=p)
    metrics = calculate_network_metrics(G)
    return {**metrics, 'p': p, 'n': n, 'k': k}

def run_experiments(n, k, p_values, num_instances):
    """
    Run multiple experiments with different rewiring probabilities
    """
    results = []
    for p in p_values:
        for _ in range(num_instances):
            result = generate_and_analyze_network(n, k, p)
            results.append(result)
    return pd.DataFrame(results)

def analyze_correlations(results_df, k_value=None):
    """
    Analyzes correlations between spectral gap and number of colors
    """
    if k_value is not None:
        k_data = results_df[results_df['k'] == k_value]
    else:
        k_data = results_df
    
    # Check if we have enough data points
    if len(k_data) < 2:
        return pd.DataFrame([{
            'k': k_value if k_value else 'all',
            'correlation': np.nan,
            'p_value': np.nan,
            'sample_size': len(k_data),
            'mean_colors': np.nan,
            'std_colors': np.nan
        }])
    
    # Check for constant values
    if k_data['spectral_gap'].std() == 0 or k_data['num_colors'].std() == 0:
        return pd.DataFrame([{
            'k': k_value if k_value else 'all',
            'correlation': np.nan,
            'p_value': np.nan,
            'sample_size': len(k_data),
            'mean_colors': k_data['num_colors'].mean(),
            'std_colors': k_data['num_colors'].std(),
            'note': 'Constant values detected - correlation undefined'
        }])
    
    # Calculate correlation if data is valid
    correlation, p_value = stats.pearsonr(
        k_data['spectral_gap'],
        k_data['num_colors']
    )
    
    return pd.DataFrame([{
        'k': k_value if k_value else 'all',
        'correlation': correlation,
        'p_value': p_value,
        'sample_size': len(k_data),
        'mean_colors': k_data['num_colors'].mean(),
        'std_colors': k_data['num_colors'].std()
    }])

def compare_to_controls(results_df): 
    """
    Compares the correlations across different network types
    """
    # Group by p value to compare across network types
    high_p = results_df[results_df['p'] > 0.4]  # Nearly random
    low_p = results_df[results_df['p'] < 0.01]  # Nearly regular
    
    results = []
    for k in results_df['k'].unique():
        normal_corr = analyze_correlations(results_df, k)
        random_corr = analyze_correlations(high_p, k)
        regular_corr = analyze_correlations(low_p, k)
        
        results.append({
            'k': k,
            'network_type': 'normal',
            **normal_corr.iloc[0].to_dict()
        })
        results.append({
            'k': k,
            'network_type': 'random',
            **random_corr.iloc[0].to_dict()
        })
        results.append({
            'k': k,
            'network_type': 'regular',
            **regular_corr.iloc[0].to_dict()
        })
    
    return pd.DataFrame(results)

def test_edge_cases():
    """
    Tests the edge cases for the network
    """
    edge_cases = {
        'high_k': {'n': 1000, 'k': 100, 'p': 0.1},
        'low_k': {'n': 1000, 'k': 4, 'p': 0.1},
        'large_n': {'n': 5000, 'k': 10, 'p': 0.1},
        'small_n': {'n': 50, 'k': 10, 'p': 0.1}
    }
    
    results = {}
    for case_name, params in edge_cases.items():
        results[case_name] = run_experiments(
            n=params['n'],
            k=params['k'],
            p_values=[params['p']],
            num_instances=30
        )
    
    return results

def save_results_to_excel(results_dict, filename='network_analysis_results'):
    """
    Save all results to Excel and plot files with error handling
    """
    try:
        # First save the plot separately
        plt.figure(figsize=(10, 6))
        plt.scatter(results_dict['results']['spectral_gap'], 
                   results_dict['results']['num_colors'], 
                   alpha=0.5)
        plt.xlabel('Spectral Gap')
        plt.ylabel('Number of Colors Used')
        plt.title('Relationship between Spectral Gap and Coloring')
        plt.grid(True)
        
        plt.yticks(range(1, 30))
        x_ticks = np.arange(0, 15.0, 1.0)
        plt.xticks(x_ticks)
        
        # Fit and plot quadratic trend line
        coefficients = np.polyfit(
            results_dict['results']['spectral_gap'], 
            results_dict['results']['num_colors'], 
            2
        )
        trendline = np.polyval(
            coefficients, 
            results_dict['results']['spectral_gap']
        )
        plt.plot(
            results_dict['results']['spectral_gap'], 
            trendline, 
            color='red', 
            label='Quadratic Trend Line'
        )
        
        plt.legend()
        
        # Save plot
        plot_filename = f"{filename}_plot.png"
        plt.savefig(plot_filename)
        plt.close()
        print(f"Plot saved to '{plot_filename}'")

        # Now save Excel data sheet by sheet
        excel_filename = f"{filename}.xlsx"
        
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Save main results
        ws = wb.create_sheet('Main Results')
        df_to_excel(results_dict['results'], ws)
        
        # Save correlation analysis
        ws = wb.create_sheet('Correlation Analysis')
        df_to_excel(results_dict['correlation_analysis'], ws)
        
        # Save control comparison
        ws = wb.create_sheet('Control Comparison')
        df_to_excel(results_dict['control_comparison'], ws)
        
        # Save edge case results
        ws = wb.create_sheet('Edge Cases')
        row = 1
        for case_name, results in results_dict['edge_case_results'].items():
            ws.cell(row=row, column=1, value=f"{case_name.upper()}:")
            row += 1
            description = results.describe()
            if 'sample_size' in description.columns:
                description = description.drop('sample_size', axis=1)
            df_to_excel(description, ws, start_row=row)
            row += len(description) + 2
        
        # Save the workbook
        wb.save(excel_filename)
        print(f"Results saved to '{excel_filename}'")
        
    except Exception as e:
        print(f"Error saving results: {str(e)}")
        raise

def df_to_excel(df, worksheet, start_row=1):
    """
    Helper function to write a DataFrame to an Excel worksheet
    """
    # Write headers
    for col_idx, column in enumerate(df.columns, 1):
        worksheet.cell(row=start_row, column=col_idx, value=str(column))
    
    # Write data
    for row_idx, row in enumerate(df.values, start_row + 1):
        for col_idx, value in enumerate(row, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

def main():
    # Graph parameters
    n = 1000  # number of nodes
    k = 120     # initial degree
    p_values = [0.001, 0.01, 0.1, 0.2, 0.3, 0.4, 0.5]
    num_instances = 10  # number of instances per configuration

    print("Running experiments...")

    # Run experiments
    results_df = run_experiments(n, k, p_values, num_instances)
    
    # Analyze results
    correlation_analysis = analyze_correlations(results_df)
    control_comparison = compare_to_controls(results_df)
    edge_case_results = test_edge_cases()
    
    # Collect all results
    results_dict = {
        'results': results_df,
        'correlation_analysis': correlation_analysis,
        'control_comparison': control_comparison,
        'edge_case_results': edge_case_results
    }
    
    # Save everything to Excel file
    filename = 'network_analysis_results29'
    save_results_to_excel(results_dict, filename) 
    
    print(f"\nResults have been saved")
    return results_dict

if __name__ == "__main__":
    results = main()